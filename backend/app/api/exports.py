"""
报价单导出API
"""
from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import HTMLResponse, StreamingResponse, FileResponse
from sqlalchemy.orm import Session
from io import BytesIO
from datetime import datetime
import json

from app.database import get_db
from app.models.quote import Quote
from app.models.quoter import Quoter
from app.models.user import User
from app.api.auth import get_current_user

router = APIRouter(prefix="/api/v1/exports", tags=["导出功能"])


def render_quote_html(quote: Quote, quoter: Quoter, theme: str = "blue") -> str:
    """
    渲染报价单HTML

    Args:
        quote: 报价单对象
        quoter: 报价人对象
        theme: 主题 (blue/gray/beige)

    Returns:
        HTML字符串
    """
    # 主题颜色配置
    themes = {
        "blue": {"primary": "#0066CC", "secondary": "#E6F2FF"},
        "gray": {"primary": "#666666", "secondary": "#F5F5F5"},
        "beige": {"primary": "#8B7355", "secondary": "#FFF8DC"}
    }

    colors = themes.get(theme, themes["blue"])

    # 构建价格表HTML
    price_table_html = ""
    if quote.template_type == "TONGPIAO":
        # 通票模式 - 区域价格表
        price_table_html = "<table class='price-table'>"
        price_table_html += "<thead><tr><th>区域</th><th>省份</th><th>首重(元/KG)</th><th>续重(元/KG)</th><th>时效</th></tr></thead>"
        price_table_html += "<tbody>"

        for region in quote.price_data.get("regions", []):
            provinces = "、".join(region.get("provinces", []))
            price_table_html += f"""
            <tr>
                <td>{region.get('regionName', '')}</td>
                <td>{provinces}</td>
                <td>{region.get('firstWeight', 0):.2f}</td>
                <td>{region.get('additionalWeight', 0):.2f}</td>
                <td>{region.get('remark', '')}</td>
            </tr>
            """

        price_table_html += "</tbody></table>"

    # 构建条款HTML
    terms_html = ""

    # 固定条款
    if quote.fixed_terms:
        terms_html += "<h3>服务条款</h3>"
        for term in quote.fixed_terms:
            terms_html += f"<p><strong>{term.get('title', '')}:</strong> {term.get('content', '')}</p>"

    # 非固定条款
    if quote.optional_terms:
        for term in quote.optional_terms:
            terms_html += f"<p><strong>{term.get('title', '')}:</strong> {term.get('content', '')}</p>"

    # 自定义条款
    if quote.custom_terms:
        terms_html += "<h3>特别说明</h3>"
        for term in quote.custom_terms:
            terms_html += f"<p>• {term}</p>"

    # HTML模板
    html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <title>报价单 - {quote.quote_number}</title>
        <style>
            * {{ margin: 0; padding: 0; box-sizing: border-box; }}
            body {{
                font-family: "Microsoft YaHei", Arial, sans-serif;
                line-height: 1.6;
                padding: 40px;
                background: #f5f5f5;
            }}
            .container {{
                max-width: 900px;
                margin: 0 auto;
                background: white;
                padding: 40px;
                box-shadow: 0 0 20px rgba(0,0,0,0.1);
            }}
            .header {{
                border-bottom: 3px solid {colors['primary']};
                padding-bottom: 20px;
                margin-bottom: 30px;
            }}
            .header h1 {{
                color: {colors['primary']};
                font-size: 28px;
                margin-bottom: 10px;
            }}
            .header .quote-number {{
                color: #666;
                font-size: 14px;
            }}
            .info-grid {{
                display: grid;
                grid-template-columns: 1fr 1fr;
                gap: 15px;
                margin-bottom: 30px;
            }}
            .info-item {{
                padding: 10px;
                background: {colors['secondary']};
                border-left: 3px solid {colors['primary']};
            }}
            .info-item label {{
                color: #666;
                font-size: 12px;
                display: block;
            }}
            .info-item value {{
                color: #333;
                font-size: 14px;
                font-weight: bold;
            }}
            .price-section {{
                margin: 30px 0;
            }}
            .price-section h2 {{
                color: {colors['primary']};
                font-size: 20px;
                margin-bottom: 15px;
                padding-bottom: 10px;
                border-bottom: 2px solid {colors['secondary']};
            }}
            .price-table {{
                width: 100%;
                border-collapse: collapse;
                margin-bottom: 20px;
            }}
            .price-table th {{
                background: {colors['primary']};
                color: white;
                padding: 12px;
                text-align: left;
                font-weight: normal;
            }}
            .price-table td {{
                padding: 10px 12px;
                border-bottom: 1px solid #eee;
            }}
            .price-table tr:hover {{
                background: {colors['secondary']};
            }}
            .terms-section {{
                margin: 30px 0;
            }}
            .terms-section h3 {{
                color: {colors['primary']};
                font-size: 16px;
                margin-bottom: 10px;
            }}
            .terms-section p {{
                margin: 8px 0;
                color: #333;
                font-size: 14px;
            }}
            .footer {{
                margin-top: 40px;
                padding-top: 20px;
                border-top: 1px solid #ddd;
                text-align: center;
                color: #999;
                font-size: 12px;
            }}
            @media print {{
                body {{ padding: 0; background: white; }}
                .container {{ box-shadow: none; }}
            }}
        </style>
    </head>
    <body>
        <div class="container">
            <div class="header">
                <h1>中通快递服务报价单</h1>
                <div class="quote-number">编号: {quote.quote_number} | 生成日期: {datetime.now().strftime('%Y-%m-%d %H:%M')}</div>
            </div>

            <div class="info-grid">
                <div class="info-item">
                    <label>客户名称</label>
                    <value>{quote.customer_name}</value>
                </div>
                <div class="info-item">
                    <label>联系人</label>
                    <value>{quote.contact_person}</value>
                </div>
                <div class="info-item">
                    <label>联系电话</label>
                    <value>{quote.contact_phone}</value>
                </div>
                <div class="info-item">
                    <label>报价日期</label>
                    <value>{quote.quote_date}</value>
                </div>
                <div class="info-item">
                    <label>有效期至</label>
                    <value>{quote.expire_date}</value>
                </div>
                <div class="info-item">
                    <label>报价人</label>
                    <value>{quoter.name} ({quoter.phone})</value>
                </div>
            </div>

            <div class="price-section">
                <h2>价格方案</h2>
                {price_table_html}
                <p style="margin-top: 10px; color: #666; font-size: 14px;">
                    {'含税价格' if quote.is_tax_included else '不含税价格'}
                </p>
            </div>

            <div class="terms-section">
                {terms_html}
            </div>

            {f'<div style="margin-top: 20px; padding: 15px; background: #fff3cd; border-left: 4px solid #ffc107;"><strong>备注:</strong> {quote.remark}</div>' if quote.remark else ''}

            <div class="footer">
                <p>中通快递服务有限公司</p>
                <p>本报价单由系统自动生成 | 如有疑问请联系报价人</p>
            </div>
        </div>
    </body>
    </html>
    """

    return html


@router.get("/quotes/{quote_id}/preview")
def preview_quote(
    quote_id: int,
    theme: str = "blue",
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    预览报价单HTML
    """
    # 查询报价单
    quote = db.query(Quote).filter(Quote.id == quote_id).first()
    if not quote:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="报价单不存在"
        )

    # 查询报价人
    quoter = db.query(Quoter).filter(Quoter.id == quote.quoter_id).first()
    if not quoter:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="报价人不存在"
        )

    # 渲染HTML
    html_content = render_quote_html(quote, quoter, theme)

    return HTMLResponse(content=html_content)


@router.get("/quotes/{quote_id}/export/html")
def export_html(
    quote_id: int,
    theme: str = "blue",
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    导出HTML文件
    """
    # 查询报价单
    quote = db.query(Quote).filter(Quote.id == quote_id).first()
    if not quote:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="报价单不存在"
        )

    # 查询报价人
    quoter = db.query(Quoter).filter(Quoter.id == quote.quoter_id).first()
    if not quoter:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="报价人不存在"
        )

    # 渲染HTML
    html_content = render_quote_html(quote, quoter, theme)

    # 返回文件
    return StreamingResponse(
        BytesIO(html_content.encode('utf-8')),
        media_type="text/html",
        headers={
            "Content-Disposition": f"attachment; filename=quote-{quote.quote_number}.html"
        }
    )


@router.get("/quotes/{quote_id}/export/excel")
def export_excel(
    quote_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    导出Excel文件
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Excel导出功能未安装"
        )

    # 查询报价单
    quote = db.query(Quote).filter(Quote.id == quote_id).first()
    if not quote:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="报价单不存在"
        )

    # 查询报价人
    quoter = db.query(Quoter).filter(Quoter.id == quote.quoter_id).first()

    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "报价单"

    # 标题
    ws['A1'] = f"报价单编号: {quote.quote_number}"
    ws['A1'].font = Font(size=14, bold=True)

    # 基本信息
    row = 3
    ws[f'A{row}'] = "客户名称:"
    ws[f'B{row}'] = quote.customer_name
    row += 1
    ws[f'A{row}'] = "联系人:"
    ws[f'B{row}'] = quote.contact_person
    row += 1
    ws[f'A{row}'] = "联系电话:"
    ws[f'B{row}'] = quote.contact_phone
    row += 1
    ws[f'A{row}'] = "报价日期:"
    ws[f'B{row}'] = str(quote.quote_date)
    row += 1
    ws[f'A{row}'] = "有效期至:"
    ws[f'B{row}'] = str(quote.expire_date)
    row += 1
    ws[f'A{row}'] = "报价人:"
    ws[f'B{row}'] = f"{quoter.name} ({quoter.phone})" if quoter else ""

    # 价格表
    row += 2
    ws[f'A{row}'] = "价格明细"
    ws[f'A{row}'].font = Font(size=12, bold=True)

    row += 1
    if quote.template_type == "TONGPIAO":
        # 表头
        headers = ['区域', '省份', '首重(元/KG)', '续重(元/KG)', '时效']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        # 数据
        for region in quote.price_data.get("regions", []):
            row += 1
            ws.cell(row=row, column=1, value=region.get('regionName', ''))
            ws.cell(row=row, column=2, value='、'.join(region.get('provinces', [])))
            ws.cell(row=row, column=3, value=region.get('firstWeight', 0))
            ws.cell(row=row, column=4, value=region.get('additionalWeight', 0))
            ws.cell(row=row, column=5, value=region.get('remark', ''))

    # 调整列宽
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 20

    # 保存到内存
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    # 返回文件
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=quote-{quote.quote_number}.xlsx"
        }
    )


@router.get("/quotes/{quote_id}/export/pdf")
def export_pdf(
    quote_id: int,
    theme: str = "blue",
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    导出PDF文件
    注意: 需要在Docker环境中使用WeasyPrint生成真实PDF
    当前返回HTML预览
    """
    # 查询报价单
    quote = db.query(Quote).filter(Quote.id == quote_id).first()
    if not quote:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="报价单不存在"
        )

    # 查询报价人
    quoter = db.query(Quoter).filter(Quoter.id == quote.quoter_id).first()
    if not quoter:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="报价人不存在"
        )

    # 渲染HTML
    html_content = render_quote_html(quote, quoter, theme)

    # TODO: 在Docker环境中使用WeasyPrint转换为PDF
    # from weasyprint import HTML
    # pdf_bytes = HTML(string=html_content).write_pdf()
    # return StreamingResponse(BytesIO(pdf_bytes), media_type="application/pdf")

    # 当前返回HTML（浏览器可以打印为PDF）
    return StreamingResponse(
        BytesIO(html_content.encode('utf-8')),
        media_type="text/html",
        headers={
            "Content-Disposition": f"inline; filename=quote-{quote.quote_number}.html"
        }
    )
